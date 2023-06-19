#!/usr/bin/env python
# coding: utf-8

# In[1]:
#!pip install openpyxl

import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
#import nltk
from nltk.corpus import stopwords
import pyttsx3

# Load the questions Excel file
questions_filename = 'questions.xlsx'
questions_workbook = load_workbook(questions_filename)
questions_sheet = questions_workbook.active

# Get the value of cell B1 from the questions Excel file
cell_b1_value = questions_sheet['B1'].value
answers_filename = cell_b1_value + '.xlsx' if cell_b1_value else 'answers.xlsx'

# Set up the Excel workbook for answers
answers_workbook = Workbook()
answers_sheet = answers_workbook.active
answers_sheet['A1'] = 'Question'
answers_sheet['B1'] = 'Answer'
row = 2  # Starting row to write answers

# Initialize NLTK and download the required resources
nltk.download('stopwords')
stop_words = set(stopwords.words('spanish'))

# Initialize pyttsx3 for text-to-speech
engine = pyttsx3.init()

# Define a function to ask questions and save answers
def ask_question(question):
    global row
    answer = input(question + " ")
    answers_sheet.cell(row=row, column=1).value = question
    answers_sheet.cell(row=row, column=2).value = answer
    row += 1  # Increment the row for the next answer

# Read the questions from the Excel file and ask them one by one
for row_data in questions_sheet.iter_rows(values_only=True):
    question = row_data[0]
    engine.say(question)  # Speak the question
    engine.runAndWait()
    ask_question(question)

# Save the answers to the specified Excel file
answers_workbook.save(answers_filename)

# Rename the answers file with the value of cell B1
if cell_b1_value and answers_filename != 'answers.xlsx':
    import os
    os.rename('answers.xlsx', answers_filename)
    
import openpyxl
import os

# Load the Excel file
filename = 'answers.xlsx'
workbook = openpyxl.load_workbook(filename)
sheet = workbook.active

# Get the value from a specific cell
cell_value = sheet['B2'].value

# Rename the file with the cell value
if cell_value:
    new_filename = f"{cell_value}.xlsx"
    os.rename(filename, new_filename)
    print(f"File renamed to: {new_filename}")
else:
    print("Cell B2 is empty. File not renamed.")


# In[ ]:




